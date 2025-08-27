import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

def saat_araliginda_mi(alarm_zamani, araliklar):
    if not isinstance(alarm_zamani, (pd.Timestamp, str)):
        return False
    try:
        alarm_saati = pd.to_datetime(alarm_zamani).time()
    except:
        return False
    for baslangic_str, bitis_str in araliklar:
        baslangic_saati = pd.to_datetime(baslangic_str).time()
        bitis_saati = pd.to_datetime(bitis_str).time()
        if baslangic_saati > bitis_saati:
            if alarm_saati >= baslangic_saati or alarm_saati <= bitis_saati:
                return True
        else:
            if baslangic_saati <= alarm_saati <= bitis_saati:
                return True
    return False

st.title("📊 Rapor İşleme Uygulaması")
rapor_file = st.file_uploader("Alarm rapor dosyasını yükleyin (.xls veya .xlsx)", type=["xls","xlsx"])
cihaz_file = st.file_uploader("Cihaz bilgileri dosyasını yükleyin (.xls veya .xlsx)", type=["xls","xlsx"])
output_file_name = st.text_input("Oluşturulacak dosya adı (örn: vardiya)", "")
saat_araliklari_str = st.text_input("Saat aralıklarını girin (örn: 05.00-08.20, 16.40-20.40)", "")

if st.button("📌 Raporu İşle"):
    if rapor_file and cihaz_file and saat_araliklari_str:
        try:
            araliklar_liste = saat_araliklari_str.replace(" ", ",").split(',')
            saat_araliklari = [tuple(aralik.strip().replace('.', ':').split('-')) for aralik in araliklar_liste if aralik.strip()]
        except:
            st.error("❌ Hatalı saat aralığı formatı!")
            st.stop()
        if not saat_araliklari:
            st.error("❌ Geçerli bir saat aralığı girilmedi.")
            st.stop()

        rapor_df = pd.read_excel(rapor_file, engine=None)
        cihaz_df = pd.read_excel(cihaz_file, engine=None)
        rapor_df = rapor_df[rapor_df['Alarm Time'].apply(lambda x: saat_araliginda_mi(x, saat_araliklari))].copy()

        column_renames = {"Vehicle":"Araç","Alarm Type":"Alarm Türü","Alarm Time":"Alarm Zamanı","Speed":"Hız","Location":"Konum"}
        value_renames = {"Camera Blocked Alarm":"Kamera Engellendi Alarmı","Making/Answering Phone Call Alarm":"Yolda Telefon Görüşmesi Yapma/Uyarma Alarmı","Fatigue Driving Alarm":"Yorgun Sürüş Alarmı","No Driver Alarm":"Kamera Engellendi Alarmı","No Fasten Seat Belt Alarm":"Emniyet Kemeri Bağlanmadı Uyarısı","Smoking Alarm":"Sigara İçme Alarmı","Driver yawning warning":"Aşırı Esneme Tespiti Alarmı","Safety Distance Alarm":"Güvenli Mesafe Alarmı","Lane Departure Alarm":"Şerit Takip / İhlal Alarmı","Forward Collision Alarm":"İleri Çarpışma Riski Alarmı"}
        drop_columns = ["Alarm Level","Altitude","Processing status","Status"]
        rapor_df.columns = [col.strip() for col in rapor_df.columns]
        rapor_df = rapor_df.rename(columns={col: column_renames[col] for col in rapor_df.columns if col in column_renames})
        rapor_df = rapor_df.drop(columns=[col for col in rapor_df.columns if col in drop_columns], errors='ignore')
        rapor_df["Alarm Türü"] = rapor_df["Alarm Türü"].replace(value_renames)
        rapor_df["Alarm Zamanı"] = pd.to_datetime(rapor_df["Alarm Zamanı"])

        cihaz_df_renamed = cihaz_df[["Plaka","Cihaz No."]].rename(columns={"Plaka":"Araç","Cihaz No.":"Cihaz No"})
        df = rapor_df.merge(cihaz_df_renamed, on="Araç", how="left")

        uyari_tr = "Sürücü Esneme Uyarısı"
        df_yawn = df[df["Alarm Türü"]==uyari_tr].copy()
        df_others = df[df["Alarm Türü"]!=uyari_tr].copy()
        yaw_keep = []
        for arac, group in df_yawn.groupby("Araç"):
            group = group.sort_values("Alarm Zamanı").reset_index(drop=True)
            i=0
            while i<len(group):
                current_time = group.loc[i,"Alarm Zamanı"]
                one_hour_later = current_time + pd.Timedelta(hours=1)
                time_window_group = group[(group["Alarm Zamanı"]>=current_time)&(group["Alarm Zamanı"]<one_hour_later)]
                if len(time_window_group)>=9:
                    yaw_keep.append(group.loc[i])
                    i=group.index.get_loc(time_window_group.index[-1])+1
                else:
                    i+=1
        df_yawn_filtered = pd.DataFrame(yaw_keep)
        df_filtered = pd.concat([df_others, df_yawn_filtered], ignore_index=True)

        filtered_rows=[]
        for (arac, alarm), group in df_filtered.groupby(["Araç","Alarm Türü"]):
            if alarm==uyari_tr:
                filtered_rows.append(group)
                continue
            group=group.sort_values("Alarm Zamanı")
            keep=[]
            last_time=None
            for _, row in group.iterrows():
                if last_time is None or (row["Alarm Zamanı"]-last_time).total_seconds()>10800:
                    keep.append(row)
                    last_time=row["Alarm Zamanı"]
            filtered_rows.append(pd.DataFrame(keep))
        df_filtered=pd.concat(filtered_rows).reset_index(drop=True)

        df_filtered["Tarih"]=df_filtered["Alarm Zamanı"].dt.date
        filtered_final=[]
        for (arac, alarm, tarih), group in df_filtered.groupby(["Araç","Alarm Türü","Tarih"]):
            if alarm==uyari_tr:
                filtered_final.append(group)
            else:
                filtered_final.append(group.head(3))
        df_final=pd.concat(filtered_final).drop(columns=["Tarih"])
        df_final["Alarm Zamanı"]=df_final["Alarm Zamanı"].dt.strftime('%Y-%m-%d %H:%M:%S')

        azaman=df_final["Alarm Zamanı"].str.replace("[- :]", "", regex=True)
        azman1=df_final["Alarm Zamanı"].str[:10]
        link_kok="\\\\10.0.0.220\\gStorage\\PERIPHERAL_FILE\\"
        df_final["Görüntü"]=[f'=HYPERLINK("{link_kok}{az1}\\0x65\\{cihaz}\\{cihaz}-{az}-0P","Görüntüye Git")' for az1, az, cihaz in zip(azman1, azaman, df_final["Cihaz No"].astype(str))]

        final_path=f"{output_file_name}_output.xlsx"
        df_final.to_excel(final_path,index=False)

        wb=load_workbook(final_path)
        ws=wb.active
        header_fill=PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")
        bold_font=Font(bold=True,size=11)
        underline_font=Font(underline="single",color="0563C1")
        center_alignment=Alignment(horizontal="center",vertical="center")
        thin=Side(border_style="thin",color="000000")
        border_all=Border(left=thin,right=thin,top=thin,bottom=thin)

        column_widths={"A":12,"B":40.11,"C":18,"D":8.89,"E":18,"F":11,"G":20}
        for col in range(1,ws.max_column+1):
            cell=ws.cell(row=1,column=col)
            cell.fill=header_fill
            cell.font=bold_font
            cell.alignment=center_alignment
            cell.border=border_all
            col_letter=get_column_letter(col)
            if col_letter in column_widths:
                ws.column_dimensions[col_letter].width=column_widths[col_letter]

        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            for cell in row:
                cell.border=border_all
                cell.alignment=center_alignment

        for row in range(2,ws.max_row+1):
            ws.cell(row=row,column=7).font=underline_font

        data=list(ws.iter_rows(min_row=2,values_only=True))
        data_sorted=sorted(data,key=lambda x:x[1])
        for i,row_data in enumerate(data_sorted,start=2):
            for j,value in enumerate(row_data,start=1):
                ws.cell(row=i,column=j).value=value

        wb.save(final_path)
        with open(final_path,"rb") as f:
            st.download_button("📥 Excel İndir",f,file_name=final_path)
    else:
        st.error("❌ Lütfen tüm alanları doldurun.")
